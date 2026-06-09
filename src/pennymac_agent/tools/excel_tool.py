"""Excel pricing-model tool (plain function for the MCP server)."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml

from ..config.settings import get_settings
from ..utils.logging import get_logger

logger = get_logger(__name__)


def _resolve_model(s, model_name: str) -> Dict[str, Any]:
    path = s.model_registry_path
    if not path.exists():
        raise FileNotFoundError(f"Model registry not found at {path}")
    models = (yaml.safe_load(path.read_text(encoding="utf-8")) or {}).get("models", {})
    if model_name not in models:
        raise KeyError(
            f"Model '{model_name}' is not registered. Known models: {list(models)}"
        )
    spec = dict(models[model_name])
    raw_path = Path(spec["path"])
    if not raw_path.is_absolute():
        raw_path = s.excel_models_dir / raw_path.name
    spec["_resolved_path"] = raw_path
    return spec


def _inputs_from_csv(data_file: str, input_map: Dict[str, str]) -> Dict[str, Any]:
    import pandas as pd

    df = pd.read_csv(data_file)
    if df.empty:
        return {}
    row = df.iloc[0].to_dict()
    return {k: row[k] for k in input_map if k in row}


def _split_ref(ref: str):
    if "!" not in ref:
        raise ValueError(f"Cell reference '{ref}' must be 'Sheet!Cell' form.")
    sheet, cell = ref.split("!", 1)
    return sheet, cell


def _run_xlwings(s, wb_path, input_map, inputs, output_map, wanted):
    import xlwings as xw

    app = xw.App(visible=s.excel_visible, add_book=False)
    try:
        book = app.books.open(str(wb_path))
        for key, value in inputs.items():
            if key not in input_map:
                raise KeyError(f"Input '{key}' not mapped for this model.")
            book.range(input_map[key]).value = value
        app.calculate()
        results: Dict[str, Any] = {}
        for key in wanted:
            if key not in output_map:
                raise KeyError(f"Output '{key}' not mapped for this model.")
            results[key] = book.range(output_map[key]).value
        book.save()
        return results
    finally:
        app.quit()


def _run_openpyxl(wb_path, input_map, inputs, output_map, wanted):
    from openpyxl import load_workbook

    wb = load_workbook(wb_path, data_only=False)
    for key, value in inputs.items():
        if key not in input_map:
            raise KeyError(f"Input '{key}' not mapped for this model.")
        sheet_name, cell = _split_ref(input_map[key])
        wb[sheet_name][cell] = value
    wb.save(wb_path)

    cached = load_workbook(wb_path, data_only=True)
    results: Dict[str, Any] = {}
    for key in wanted:
        if key not in output_map:
            raise KeyError(f"Output '{key}' not mapped for this model.")
        sheet_name, cell = _split_ref(output_map[key])
        results[key] = cached[sheet_name][cell].value
    return results


def excel_model(
    model_name: str,
    inputs: Optional[Dict[str, Any]] = None,
    outputs: Optional[List[str]] = None,
    data_file: Optional[str] = None,
    engine: str = "auto",
) -> str:
    """Drive a registered Excel pricing model: write inputs, recalc, read outputs.

    Reference the model by its logical name from models/registry.yaml. Optionally
    load inputs from a CSV produced by a data tool via 'data_file' (the first
    row's columns matching the model's input keys are used).

    Args:
        model_name: Logical model key from models/registry.yaml.
        inputs: Map of friendly input keys to values, e.g. {"coupon": 5.5}.
        outputs: Friendly output keys to read back; omit to read all.
        data_file: Optional CSV path to source inputs from (first row).
        engine: 'auto' (prefer xlwings), 'xlwings', or 'openpyxl'.
    """
    s = get_settings()
    try:
        spec = _resolve_model(s, model_name)
    except (FileNotFoundError, KeyError) as exc:
        return f"Error: {exc}"

    wb_path: Path = spec["_resolved_path"]
    input_map: Dict[str, str] = spec.get("inputs", {})
    output_map: Dict[str, str] = spec.get("outputs", {})
    wanted = outputs if outputs is not None else list(output_map)

    if inputs is None and data_file:
        try:
            inputs = _inputs_from_csv(data_file, input_map)
        except Exception as exc:
            return f"Error: could not read data_file '{data_file}': {exc}"
    inputs = inputs or {}

    if not wb_path.exists():
        return (
            f"Error: workbook for model '{model_name}' not found at {wb_path}. "
            "Place the file there or update its registry path."
        )

    use_xlwings = engine in ("auto", "xlwings")
    try:
        if use_xlwings:
            try:
                results = _run_xlwings(s, wb_path, input_map, inputs, output_map, wanted)
            except ImportError:
                if engine == "xlwings":
                    raise
                logger.warning("xlwings unavailable; falling back to openpyxl")
                results = _run_openpyxl(wb_path, input_map, inputs, output_map, wanted)
        else:
            results = _run_openpyxl(wb_path, input_map, inputs, output_map, wanted)
    except Exception as exc:  # pragma: no cover - depends on local Excel
        logger.exception("Excel model run failed")
        return f"Error: Excel run failed: {exc}"

    return (
        f"Model '{model_name}': wrote {len(inputs)} input(s), read "
        f"{len(results)} output(s) -> {results}"
    )
