"""Excel Modeling Agent.

Pushes inputs into complex pricing-model workbooks and extracts calculated
results. Uses xlwings when Excel is available (so formulas recalc live) and
falls back to openpyxl for headless read/write.

Workbooks and their input/output cell mappings are declared in
``models/registry.yaml`` (see WORKFLOWS.md, section B) and looked up by a
logical model name.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml

from ..config.settings import Settings, get_settings
from ..utils.logging import get_logger
from .base_agent import AgentResult, BaseAgent

logger = get_logger(__name__)


class ExcelModelingAgent(BaseAgent):
    name = "excel_model"
    description = (
        "Drive a complex Excel pricing-model workbook. Use this to push input "
        "values (e.g. coupon, UPB, settle date) into a registered model, "
        "trigger recalculation, and read back calculated outputs (e.g. price, "
        "OAS, duration). Reference the model by its logical name from "
        "models/registry.yaml."
    )
    parameters = {
        "type": "object",
        "properties": {
            "model_name": {
                "type": "string",
                "description": "Logical model key from models/registry.yaml.",
            },
            "inputs": {
                "type": "object",
                "description": (
                    "Map of friendly input keys (from the registry) to values "
                    "to write, e.g. {'coupon': 5.5, 'upb': 1000000}."
                ),
            },
            "outputs": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    "Friendly output keys to read back. Omit to return all "
                    "outputs declared for the model."
                ),
            },
            "engine": {
                "type": "string",
                "enum": ["auto", "xlwings", "openpyxl"],
                "description": "Calc engine. 'auto' prefers xlwings.",
                "default": "auto",
            },
        },
        "required": ["model_name"],
        "additionalProperties": False,
    }

    def __init__(self, settings: Optional[Settings] = None):
        self.settings = settings or get_settings()

    # --- registry -------------------------------------------------------------
    def _load_registry(self) -> Dict[str, Any]:
        path = self.settings.model_registry_path
        if not path.exists():
            raise FileNotFoundError(f"Model registry not found at {path}")
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        return data.get("models", {})

    def _resolve_model(self, model_name: str) -> Dict[str, Any]:
        models = self._load_registry()
        if model_name not in models:
            raise KeyError(
                f"Model '{model_name}' is not registered. Known models: "
                f"{list(models)}"
            )
        spec = models[model_name]
        raw_path = Path(spec["path"])
        if not raw_path.is_absolute():
            raw_path = self.settings.excel_models_dir / raw_path.name
        spec = dict(spec)
        spec["_resolved_path"] = raw_path
        return spec

    # --- engines --------------------------------------------------------------
    def _run_xlwings(
        self,
        wb_path: Path,
        input_map: Dict[str, str],
        inputs: Dict[str, Any],
        output_map: Dict[str, str],
        wanted: List[str],
    ) -> Dict[str, Any]:
        import xlwings as xw

        app = xw.App(visible=self.settings.excel_visible, add_book=False)
        try:
            book = app.books.open(str(wb_path))
            for key, value in inputs.items():
                if key not in input_map:
                    raise KeyError(f"Input '{key}' not mapped for this model.")
                book.sheets.active  # ensure book active
                book.range(input_map[key]).value = value
            app.calculate()
            results: Dict[str, Any] = {}
            for key in wanted:
                if key not in output_map:
                    raise KeyError(f"Output '{key}' not mapped for this model.")
                results[key] = book.range(output_map[key]).value
            book.save()
            return results
        finally:
            app.quit()

    def _run_openpyxl(
        self,
        wb_path: Path,
        input_map: Dict[str, str],
        inputs: Dict[str, Any],
        output_map: Dict[str, str],
        wanted: List[str],
    ) -> Dict[str, Any]:
        # NOTE: openpyxl does not recalculate formulas. Reading outputs returns
        # the last value cached by Excel. Use the xlwings engine for live recalc.
        from openpyxl import load_workbook

        wb = load_workbook(wb_path, data_only=False)
        for key, value in inputs.items():
            if key not in input_map:
                raise KeyError(f"Input '{key}' not mapped for this model.")
            sheet_name, cell = _split_ref(input_map[key])
            wb[sheet_name][cell] = value
        wb.save(wb_path)

        cached = load_workbook(wb_path, data_only=True)
        results: Dict[str, Any] = {}
        for key in wanted:
            if key not in output_map:
                raise KeyError(f"Output '{key}' not mapped for this model.")
            sheet_name, cell = _split_ref(output_map[key])
            results[key] = cached[sheet_name][cell].value
        return results

    # --- main entrypoint ------------------------------------------------------
    def run(  # type: ignore[override]
        self,
        model_name: str,
        inputs: Optional[Dict[str, Any]] = None,
        outputs: Optional[List[str]] = None,
        engine: str = "auto",
        **_: Any,
    ) -> AgentResult:
        try:
            spec = self._resolve_model(model_name)
        except (FileNotFoundError, KeyError) as exc:
            return AgentResult.failure(str(exc))

        wb_path: Path = spec["_resolved_path"]
        input_map: Dict[str, str] = spec.get("inputs", {})
        output_map: Dict[str, str] = spec.get("outputs", {})
        wanted = outputs if outputs is not None else list(output_map)
        inputs = inputs or {}

        if not wb_path.exists():
            return AgentResult.failure(
                f"Workbook for model '{model_name}' not found at {wb_path}. "
                "Place the file there or update its path in the registry."
            )

        use_xlwings = engine in ("auto", "xlwings")
        try:
            if use_xlwings:
                try:
                    results = self._run_xlwings(
                        wb_path, input_map, inputs, output_map, wanted
                    )
                except ImportError:
                    if engine == "xlwings":
                        raise
                    logger.warning("xlwings unavailable; falling back to openpyxl")
                    results = self._run_openpyxl(
                        wb_path, input_map, inputs, output_map, wanted
                    )
            else:
                results = self._run_openpyxl(
                    wb_path, input_map, inputs, output_map, wanted
                )
        except Exception as exc:  # pragma: no cover - depends on local Excel
            logger.exception("Excel model run failed")
            return AgentResult.failure(f"Excel run failed: {exc}")

        summary = (
            f"Model '{model_name}': wrote {len(inputs)} input(s), read "
            f"{len(results)} output(s) -> {results}"
        )
        return AgentResult.success(summary, data=results, model=model_name)


def _split_ref(ref: str) -> tuple[str, str]:
    """Split a 'Sheet!A1' reference into (sheet, cell)."""
    if "!" not in ref:
        raise ValueError(f"Cell reference '{ref}' must be 'Sheet!Cell' form.")
    sheet, cell = ref.split("!", 1)
    return sheet, cell
